# Python tool to export and import excel files to and from the json

import pandas as pd
import os
import json

from load_schema_json import load_json, validate_json, json_to_dataframe
from openpyxl.styles import PatternFill, Border, Side, Font

def excel_to_json(excel_filename, json_filename_out):
    """
    Convert the Excel file to JSON data.

    Parameters
    ----------
    excel_filename : str
        Input Excel filename
    json_filename : str
        JSON filename
    """
    df = pd.read_excel(excel_filename)
    data_dict = dataframe_to_json(df)

    # write dict to json
    data_json = json.dumps(data_dict, indent=2)

    # replace NaN with None
    data_json = data_json.replace('NaN', 'null')

    # write json to file
    with open(json_filename_out, 'w') as f:
        f.write(data_json)

     # validate the data
    json_schema = "../schemas/schema_sequencing_examples_reason.json"
    schema = load_json(json_schema)
    data_loaded = load_json(json_filename_out)
    assert validate_json(data_loaded, schema), "JSON data is invalid!"



def dataframe_to_json(df):
    """
    converts dataframe to json

    Parameters
    ----------
    df : pandas.DataFrame
        dataframe to convert

    Returns
    -------
    data : dict
        JSON data
    """
    data = {}
    data['Data'] = []
    # group by type, subtype, subsubtype
    df = df.groupby(['Type', 'Subtype', 'Sub_Subtype'])
    # iterate through the groups
    for name, group in df:
        item = {}
        item['Type'] = name[0]
        item['Subtype'] = name[1]
        item['Sub_Subtype'] = name[2]
        item['Examples'] = []
        # iterate through the rows in the group
        for index, row in group.iterrows():
            example = {}
            example['Example'] = row['Example']
            example['Reasoning'] = row['Reasoning']
            example['Linkage_Word'] = row['Linkage_Word']
            item['Examples'].append(example)
        data['Data'].append(item)
    return data


def json_to_excel(
    json_filename, 
    excel_filename, 
    json_schema =  "../schemas/schema_sequencing_examples_reason.json"):
    """
    Convert the JSON data to an Excel file.

    Parameters
    ----------
    json_filename : str
        JSON filename
    excel_filename : str
        Output Excel filename
    json_schema : str
    """
    data = load_json(json_filename)
    schema = load_json(json_schema)

    assert validate_json(data, schema), "JSON data is invalid!"
    
    df = json_to_dataframe(data)
    
    #df.to_excel(excel_filename, index=False)
    write_excel_formatting(df, excel_filename)


def write_excel_formatting(df, path):
    """ Create a writer object with excel formatting

    Parameters
    ----------
    df : pandas.DataFrame
    path : str, output path for the excel file
    """

    # Create a writer object
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        
        # Access the openpyxl worksheet object
        worksheet = writer.sheets['Sheet1']
        
        # Get openpyxl objects for styles
        from openpyxl.styles import PatternFill, Font
        
        # Define header format
        header_fill = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
        header_font = Font(bold=True)
        
        # Apply formatting to headers
        for cell in worksheet["1:1"]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Set column widths (adjust as necessary)
        col_widths = {
            'A': 20,
            'B': 20,
            'C': 20,
            'D': 80,
            'E': 80,
            'F': 20
        }
        for col, width in col_widths.items():
            worksheet.column_dimensions[col].width = width



def test_json_to_excel():
    """
    Test json_to_excel().
    """
    json_filename = "../schemas/sequencing_examples_reason.json"
    excel_filename = "../schemas/sequencing_examples_reason.xlsx"
    json_to_excel(json_filename, excel_filename)


def test_excel_to_json():
    """
    Test json_to_excel().
    """
    
    excel_filename = "../schemas/sequencing_examples_reason.xlsx"
    json_filename_out = "../schemas/sequencing_examples_reason_recovered.json"
    excel_to_json(excel_filename, json_filename_out)